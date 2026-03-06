"""
CSV and Excel export service for reports and data lists.

Provides ?format=csv and ?format=xlsx for report endpoints.
Uses openpyxl for Excel generation, csv stdlib for CSV.
"""

import csv
import io
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, numbers


def _serialize_value(val: Any) -> Any:
    """Convert value for export."""
    if isinstance(val, Decimal):
        return float(val)
    if val is None:
        return ""
    return val


def rows_to_csv(headers: list[str], rows: list[list[Any]]) -> bytes:
    """Generate CSV bytes from headers and row data."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([_serialize_value(v) for v in row])
    return output.getvalue().encode("utf-8")


def rows_to_xlsx(headers: list[str], rows: list[list[Any]], sheet_name: str = "Report") -> bytes:
    """Generate XLSX bytes from headers and row data."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header styling
    header_font = Font(bold=True, size=11)
    header_border = Border(bottom=Side(style="thin"))

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_serialize_value(val))
            if isinstance(val, (Decimal, float)):
                cell.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1

    # Auto-width columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def profit_loss_to_rows(report) -> tuple[list[str], list[list[Any]]]:
    """Convert ProfitLossReport to header/rows format."""
    headers = ["Category", "Account Number", "Account Name", "Balance"]
    rows = []
    for item in report.revenue_items:
        rows.append(["Revenue", item.account_number, item.account_name, item.balance])
    rows.append(["", "", "Total Revenue", report.total_revenue])
    rows.append([])
    for item in report.expense_items:
        rows.append(["Expense", item.account_number, item.account_name, item.balance])
    rows.append(["", "", "Total Expenses", report.total_expenses])
    rows.append([])
    rows.append(["", "", "Net Income", report.net_income])
    return headers, rows


def balance_sheet_to_rows(report) -> tuple[list[str], list[list[Any]]]:
    """Convert BalanceSheetReport to header/rows format."""
    headers = ["Category", "Account Number", "Account Name", "Balance"]
    rows = []
    for item in report.assets:
        rows.append(["Asset", item.account_number, item.account_name, item.balance])
    rows.append(["", "", "Total Assets", report.total_assets])
    rows.append([])
    for item in report.liabilities:
        rows.append(["Liability", item.account_number, item.account_name, item.balance])
    rows.append(["", "", "Total Liabilities", report.total_liabilities])
    rows.append([])
    for item in report.equity:
        rows.append(["Equity", item.account_number, item.account_name, item.balance])
    rows.append(["", "", "Total Equity", report.total_equity])
    return headers, rows


def cash_flow_to_rows(report) -> tuple[list[str], list[list[Any]]]:
    """Convert CashFlowReport to header/rows format."""
    headers = ["Category", "Description", "Amount"]
    rows = []
    rows.append(["Operating", "Net Income", report.net_income])
    for item in report.operating_activities:
        rows.append(["Operating", item.description, item.amount])
    rows.append(["", "Total Operating", report.total_operating])
    rows.append([])
    for item in report.investing_activities:
        rows.append(["Investing", item.description, item.amount])
    rows.append(["", "Total Investing", report.total_investing])
    rows.append([])
    for item in report.financing_activities:
        rows.append(["Financing", item.description, item.amount])
    rows.append(["", "Total Financing", report.total_financing])
    rows.append([])
    rows.append(["", "Net Change in Cash", report.net_change])
    rows.append(["", "Beginning Cash", report.beginning_cash])
    rows.append(["", "Ending Cash", report.ending_cash])
    return headers, rows


def ar_aging_to_rows(report) -> tuple[list[str], list[list[Any]]]:
    """Convert ARAgingReport to header/rows format."""
    headers = ["Customer", "Invoice #", "Invoice Date", "Due Date", "Days Outstanding", "Current", "1-30", "31-60", "61-90", "90+", "Total"]
    rows = []
    for item in report.items:
        rows.append([
            item.customer_name, item.invoice_number, str(item.invoice_date),
            str(item.due_date), item.days_outstanding,
            item.current, item.bucket_1_30, item.bucket_31_60,
            item.bucket_61_90, item.bucket_90_plus, item.total,
        ])
    rows.append([
        "TOTAL", "", "", "", "",
        report.total_current, report.total_1_30, report.total_31_60,
        report.total_61_90, report.total_90_plus, report.grand_total,
    ])
    return headers, rows


def ap_aging_to_rows(report) -> tuple[list[str], list[list[Any]]]:
    """Convert APAgingReport to header/rows format."""
    headers = ["Vendor", "Bill #", "Bill Date", "Due Date", "Days Outstanding", "Current", "1-30", "31-60", "61-90", "90+", "Total"]
    rows = []
    for item in report.items:
        rows.append([
            item.vendor_name, item.bill_number, str(item.bill_date),
            str(item.due_date), item.days_outstanding,
            item.current, item.bucket_1_30, item.bucket_31_60,
            item.bucket_61_90, item.bucket_90_plus, item.total,
        ])
    rows.append([
        "TOTAL", "", "", "", "",
        report.total_current, report.total_1_30, report.total_31_60,
        report.total_61_90, report.total_90_plus, report.grand_total,
    ])
    return headers, rows


def generic_list_to_export(items: list[dict], columns: list[str]) -> tuple[list[str], list[list[Any]]]:
    """Convert a list of dicts to header/rows using specified column keys."""
    rows = []
    for item in items:
        rows.append([item.get(col, "") for col in columns])
    return columns, rows
